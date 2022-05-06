import os
import csv
from copy import copy
from openpyxl import *
from datetime import date
Inputs ={}
analyst_name = ''
geo_dict ={}
def Read_Config(input_csv): 
    config_inputs={}
    with open(input_csv,"r") as csv_file:
        csv_reader = csv.DictReader(csv_file,delimiter="=")
        for line in csv_reader:
            k = line["variable"]
            if k.startswith("#"): #skips comment rows in csv config file
                continue
            elif k == "Geo" or k=="GeoContacts":
                config_inputs[k] = eval("{"+line["value"]+"}")
            else:
                config_inputs[k] = line["value"]        
    return config_inputs

def MostRecentFile():
    with os.scandir() as it:
        files = {}
        for entry in it:
            if entry.is_file():
                if entry.name.endswith('xlsx'):
                    files.setdefault(entry.stat().st_mtime,entry.name)
                    #print(entry.name, entry.stat().st_ctime)
    files=(sorted(files.items(),reverse=True))
    fileName = files[0][1]
    return fileName

def CopyHeader(s_ws,s_row,t_ws):
    for row in s_ws.iter_rows(s_row,s_row):#,values_only = True):
        for cell in row:
            new_cell = t_ws.cell(row=cell.row-s_row+1, column=cell.col_idx, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
def CopyCellFormat(s_ws,s_row):
    print(" Formatting cell.. This may take a moment...")
    for row in s_ws.iter_rows(s_row,s_row):#,values_only = True):
       # print(row)
        for i in range(s_row+1,s_ws.max_row+1):
            for cell in row:
                new_cell = s_ws.cell(row=cell.row+i-s_row, column=cell.col_idx)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
    print("Formating Complete.")

def FindCellCoordinate(ws,search_value=None):
    for row in ws.values:
        try:
            x = list(row).index(search_value)
            return x
        except:
            pass
def UploadApproved(ws, destFile):
    db = load_workbook(destFile,data_only=False, read_only=False)
    ds = db['Surcharge Request']
    VndCol = FindCellCoordinate(ws, 'Vendor')
    LastCol = FindCellCoordinate(ws, 'Surcharge Desc') +1
    plantCol = FindCellCoordinate(ws,'Plant')
    upload_data = []
    for row_data in ws.iter_rows(2, min_col=VndCol, max_col=LastCol, values_only=True):
        upload_data.append(list(row_data))

    formularow = list(ds[ds.max_row])
    RowNum = formularow[0].value
    for row in upload_data:  # (0, len(upload_data)):
        RowNum += 1
        row.insert(0, RowNum)
        row.insert(1,date.today())
        row.insert(2,analyst_name)
        row.insert(3, formularow[3].value)
        try:
            geo = geo_dict[row[plantCol-VndCol+5]]
        except:
            geo = 'UNK'

        row.insert(4,geo)
        for i in range(5, VndCol - 1):
            if str(formularow[i].value).startswith("="):
                row.insert(i, formularow[i].value)
            else:
                row.insert(i,"")
        ds.append(row)
    print("Upload Complete.")
    CopyCellFormat(ds,5)
    db.save(destFile)


#wb = load_workbook(xl,data_only=True, read_only=False)
def SortSucharges(xl_file):
    print("Begin Sorting.")
    wb = load_workbook(xl_file,data_only=True, read_only=False)
    ws = wb['Surcharge Request']
    wb.create_sheet('AutoApprove')
    ws_AutoAp = wb['AutoApprove']
    CopyHeader(ws,4,ws_AutoAp)
    testCol = FindCellCoordinate(ws,'Surcharge% per PO line')
    plantCol = FindCellCoordinate(ws,'Plant')
    geoCol = FindCellCoordinate(ws,'GEO')
    buyGrCol = FindCellCoordinate(ws,'Buy Group')
    for row_data in ws.iter_rows(5,values_only=True):
        try:
            geo = geo_dict[row_data[plantCol]]
        except:
            geo = 'UNK'
        SheetList = wb.sheetnames
        row_data = list(row_data)
        row_data[geoCol] = geo
        try:
            if row_data[testCol] <= 0.5 or (row_data[testCol] <= 1 and row_data[buyGrCol] == 'PROMOTIONAL'):
                ws_AutoAp.append(row_data)
            elif geo not in SheetList:
                wb.create_sheet(geo)
                tmws = wb[geo]
                CopyHeader(ws,4,tmws)
                tmws.append(row_data)
            else:
                tmws = wb[geo]
                tmws.append(row_data)
        except:
            print('Invalid values in row')
    wb.save(xl)
print('Reading config.csv..')
Inputs = Read_Config('config.csv')
geo_dict = Inputs['Geo']
analyst_name = Inputs['analyst_name']
message = Inputs['Message']
xl=MostRecentFile()
print(message)
print("Processing file: ",xl)
SortSucharges(xl)
print("Sorting Complete. Beginning Auto approved items upload...")
wb = load_workbook(xl,data_only=True, read_only=False)
ws = wb['AutoApprove']
UploadApproved(ws,'sharepoint\Dest.xlsx')
email_contacts=[]
for name in ws.sheetnames:
    email_contacts.append(GeoContacts[name])
#TODO: email excel for geo approval
print("All done!")


